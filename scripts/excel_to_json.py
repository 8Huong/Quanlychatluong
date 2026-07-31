"""
excel_to_json.py
-----------------
Chuyển đổi file Excel "Quản lý Đề án Cải tiến Chất lượng" thành data.json / data.js
để website dashboard sử dụng.

Cách dùng:
    python3 excel_to_json.py "duong_dan_den_file.xlsx"

Kết quả:
    ../data/data.json   -> dữ liệu thô (dùng để debug / import Firebase sau này)
    ../js/data.js        -> dữ liệu nhúng thẳng vào biến JS (website đọc file này)
"""
import openpyxl, json, datetime, sys, os

def cell_val(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if v is None:
        return None
    return v

def sheet_to_records(ws, headers_map):
    headers = [c.value for c in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).strip() == '':
            continue
        rec = {}
        for h, v in zip(headers, row):
            key = headers_map.get(h, h)
            rec[key] = cell_val(v)
        records.append(rec)
    return records

KHOA_H = {
    'Mã khoa': 'ma_khoa', 'Tên khoa/phòng': 'ten_khoa',
    'Khối chuyên môn': 'khoi', 'Trưởng khoa/Phụ trách': 'truong_khoa'
}
DEAN_H = {
    'Mã đề án': 'ma_de_an', 'Năm': 'nam', 'Tên đề án': 'ten_de_an',
    'Khoa/\nphòng': 'khoa_phong', 'Lĩnh vực cải tiến': 'linh_vuc',
    'Mức ưu tiên': 'muc_uu_tien', 'Chủ nhiệm đề án': 'chu_nhiem',
    'Ngày phê duyệt': 'ngay_phe_duyet', 'Ngày bắt đầu kế hoạch': 'ngay_bat_dau',
    'Ngày kết thúc kế hoạch': 'ngay_ket_thuc',
    'Ngày hoàn thành thực tế': 'ngay_hoan_thanh_tt', 'Trạng thái': 'trang_thai_nhap'
}
MUCTIEU_H = {
    'Mã mục tiêu': 'ma_muc_tieu', 'Mã Đề án': 'ma_de_an', 'Mục tiêu': 'muc_tieu',
    'Loại mục tiêu': 'loai_muc_tieu', 'Chỉ số': 'chi_so',
    'Ngưỡng mục tiêu': 'nguong_muc_tieu', 'Đơn vị': 'don_vi',
    'Giá trị trước cải tiến': 'gia_tri_truoc', 'Giá trị sau cải tiến': 'gia_tri_sau',
    'Kết quả': 'ket_qua', 'Thời hạn hoàn thành': 'thoi_han',
    'Trạng thái mục tiêu': 'trang_thai_mt'
}
HOATDONG_H = {
    'Mã hoạt động': 'ma_hd', 'Mã đề án': 'ma_de_an', 'Hoạt động': 'ten_hd',
    'Người phụ trách': 'nguoi_phu_trach', 'Ngày bắt đầu': 'ngay_bat_dau',
    'Ngày kết thúc': 'ngay_ket_thuc', 'Trạng thái': 'trang_thai',
    '% hoàn thành hoạt động': 'phan_tram', 'Minh chứng': 'minh_chung'
}
TIENDO_H = {
    'Tháng': 'thang', 'Mã đề án': 'ma_de_an', 'Đánh giá tiến độ': 'danh_gia',
    'Khó khăn': 'kho_khan', 'Kiến nghị': 'kien_nghi',
    'Người cập nhật': 'nguoi_cap_nhat', 'Ngày cập nhật': 'ngay_cap_nhat'
}
ACTIONLOG_H = {'Ngày': 'ngay', 'Khoa': 'khoa', 'Nội dung': 'noi_dung', 'Kết quả': 'ket_qua'}

def main():
    src = sys.argv[1] if len(sys.argv) > 1 else \
        "/mnt/user-data/uploads/Quan_ly_De_an_Cai_tien_Chat_luong_BV_Tan_Phu_-_Copy.xlsx"
    wb = openpyxl.load_workbook(src, data_only=True)

    data = {
        'khoa': sheet_to_records(wb['DM_Khoa'], KHOA_H),
        'dean': sheet_to_records(wb['DeAn'], DEAN_H),
        'muctieu': sheet_to_records(wb['Muctieu'], MUCTIEU_H),
        'hoatdong': sheet_to_records(wb['HoatDong'], HOATDONG_H),
        'tiendo': sheet_to_records(wb['TienDo'], TIENDO_H),
        'actionlog': sheet_to_records(wb['Actionlog'], ACTIONLOG_H),
        'generated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    here = os.path.dirname(os.path.abspath(__file__))
    out_json = os.path.join(here, '..', 'data', 'data.json')
    out_js = os.path.join(here, '..', 'js', 'data.js')

    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    with open(out_js, 'w', encoding='utf-8') as f:
        f.write('// File này được tạo tự động bởi scripts/excel_to_json.py\n')
        f.write('// KHÔNG sửa tay - hãy sửa trong file Excel gốc rồi chạy lại script.\n')
        f.write('window.MASTER_DATA = ')
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(';\n')

    print(f"OK: {len(data['dean'])} đề án, {len(data['khoa'])} khoa/phòng, "
          f"{len(data['hoatdong'])} hoạt động, {len(data['muctieu'])} mục tiêu, "
          f"{len(data['tiendo'])} nhật ký tiến độ, {len(data['actionlog'])} hoạt động nhật ký.")
    print(f"-> {out_json}")
    print(f"-> {out_js}")

if __name__ == '__main__':
    main()
